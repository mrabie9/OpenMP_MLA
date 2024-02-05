import subprocess
import os
import time
from datetime import datetime
import numpy as np
import openpyxl as xl

data_wdir = "C:\WSD530\OpenMP/data/"
wdir = "C:\WSD530\OpenMP/codes/part1_cases/NeuralNet/"
program = "ann"
app = "ann_omp_"

# mla_programs = ["SVM", "NeuralNet", "RandomForest"]

def rebuild():
    print("Rebuilding...")
    print(wdir +  "&& make")
    compile = subprocess.run("cd " + wdir + "&& make clean", shell=True, capture_output=True, text=True)
    if compile.stderr:
        print(compile.stderr)
        exit()
    compile = subprocess.run("cd " + wdir + "&& make", shell=True, capture_output=True, text=True)
    if compile.stderr:
        print(compile.stderr)
        exit()
    print("Rebuilt.")

def update_summary_ws():
    # Update summary
    workbook = xl.load_workbook("C:/WSD530/OpenMP/data/Performance.xlsx")
    ws = workbook['MLA Summary']
    algo = ["'ANN'", "'RF'", "'SVM'"]
    #"!B5", "!E5", 
    cells = ["!H5", "!K5", "!N5", "!Q5", "!T5", "!W5", "!Z5", "!AC5", "!AF5", "!AI5", "!AL5", "!AO5", "!AR5", "!AU5", "!AX5", "!BA5", "!BD5"]
    row = 5
    for j in range(3):
        if j >0:
            cells = ["!B5", "!E5", "!H5", "!K5", "!N5", "!Q5", "!T5", "!W5", "!Z5", "!AC5", "!AF5", "!AI5", "!AL5", "!AO5", "!AR5", "!AU5", "!AX5"]
        col = 3
        for i in range(17):
            ws.cell(row=row, column=col, value = "="+algo[j]+cells[i] )
            col+=1
        row +=1  
    workbook.save('C:/WSD530/OpenMP/data/Performance.xlsx')

iterations = 100
def main():
    global app
    col = 1

    # Rebuild programs
    rebuild()

    # run base (sequential) program
    workbook = xl.load_workbook("C:/WSD530/OpenMP/data/Performance.xlsx")
    ws = workbook[program]

    # 
    opts = ["", "_O3"]   

    # original program
    for opt in opts:
        output = data_wdir + program + "/output/" + program + "_base" + opt + ".txt"
        exec_times = data_wdir + program + "/performance/" + program + "_base" + opt + ".txt"

        with open(exec_times, 'w') as file:
            file.close()
        print(wdir + "base/.\\" + program + opt)
        row = 6
        for i in range(iterations):
            log = subprocess.run(wdir + "base/.\\" + program + opt, shell=True, capture_output=True, text=True)
            # print(log.stderr)
            with open(output, 'w') as file:
                file.write(log.stdout)
                file.close()
            # print("Completed " + program )

            with open(output, 'r') as file:
                lines = file.readlines()
                file.close()
            
            if lines:
                with open(exec_times, 'a') as file:
                    file.write(lines[-1])
                    file.close()
            else:
                print("No lines at i = " + str(i) + ", base")
            
            # write to spreadsheet
            ws.cell(row=row, column=col, value = float(lines[-1][:-1]))
            row += 1
        col +=3
    workbook.save('C:/WSD530/OpenMP/data/Performance.xlsx')

    # modified sequential
    for opt in opts:
        output = data_wdir + program + "/output/" + program + "_base_nll" + opt + ".txt"
        exec_times = data_wdir + program + "/performance/" + program + "_base_nll" + opt + ".txt"

        with open(exec_times, 'w') as file:
            file.close()
        print(wdir + "non_ll/.\\" + program + "_nonll" + opt)
        row = 6
        for i in range(iterations):
            log = subprocess.run(wdir + "non_ll/.\\" + program + "_nonll" + opt, shell=True, capture_output=True, text=True)
            # print(log.stderr)
            with open(output, 'w') as file:
                file.write(log.stdout)
                file.close()
            # print("Completed " + program )

            with open(output, 'r') as file:
                lines = file.readlines()
                file.close()
            
            if lines:
                with open(exec_times, 'a') as file:
                    file.write(lines[-1])
                    file.close()
            else:
                print("No lines at i = " + str(i) + ", base")
            
            ws.cell(row=row, column=col, value = float(lines[-1][:-1]))
            row += 1
        col +=3

    workbook.save('C:/WSD530/OpenMP/data/Performance.xlsx') 
    # Update summary
    # update_summary_ws()
    
    # Vary number of threads and schedule
    th_variants = ["2_threads", "4_threads", "8_threads", "16_threads", "32_threads"]
    threads = ["2", "4", "8", "16", "32"]
    sch_variants = ["Dynamic", "Static", "Guided"]
    idx = -1
    for th in th_variants:
        idx +=1
        for sch in sch_variants:
            print("                                          ", end="\r")
            print("Thread: " + th + " - Scheduler: " + sch, end="\r")
            output = data_wdir + program + "/output/" + program + "_" + th + "_" + sch + ".txt"
            exec_times = data_wdir + program + "/performance/" + program + "_" + th + "_" + sch + ".txt"

            with open(exec_times, 'w') as file:
                file.close()

            row = 6
            for i in range(iterations):
                log = subprocess.run(wdir + th + "/" + sch + "/.\\" + app + sch + threads[idx], shell=True, capture_output=True, text=True)
                # print(log.stderr)
                with open(output, 'w') as file:
                    file.write(log.stdout)
                    file.close()
                # print("Completed " + program )

                with open(output, 'r') as file:
                    lines = file.readlines()
                    file.close()
                
                if lines:
                    with open(exec_times, 'a') as file:
                        file.write(lines[-1])
                        file.close()
                else:
                    print("No lines at i = " + str(i) + ", " + th + ", " + sch)
                ws.cell(row=row, column=col, value = float(lines[-1][:-1]))
                row += 1
            col +=3
        print("                                          ", end="\r")
        print(th + " done.")
    
    workbook.save('C:/WSD530/OpenMP/data/Performance.xlsx')
    

main()